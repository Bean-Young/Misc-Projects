def test2():
    import hashlib
    import os

    def get_md5(file_path):
        if not os.path.exists(file_path):
            return "文件不存在"
        md5 = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                md5.update(chunk)
        return md5.hexdigest()

    file_path = input("请输入文件名：")
    print(get_md5(file_path))

def test4():
    import random
    import string
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    def generate_name():
        first_name = ''.join(random.choices(string.ascii_uppercase, k=2))
        last_name = ''.join(random.choices(string.ascii_lowercase, k=3))
        return f"{first_name} {last_name}"

    def generate_score():
        return random.randint(0, 100)

    names=[]
    for i in range(10):
        names.append ( generate_name())
    data = []
    courses = ['语文', '数学', '英语','科学']
    for i in range(200):
        name=names[random.randint(0,9)]
        course=courses[random.randint(0,3)]
        score = generate_score()
        l=[]
        l.append(name)
        l.append(course)
        l.append(score)
        data.append(l)

    df = pd.DataFrame(data, columns=['姓名', '课程', '成绩'])
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save('成绩表.xlsx')

    result = df.groupby(['课程', '姓名'])['成绩'].max().reset_index()
    result.columns = ['课程', '姓名', '最高成绩']

    wb_result = Workbook()
    ws_result = wb_result.active
    for r in dataframe_to_rows(result, index=False, header=True):
        ws_result.append(r)
    for i in range(4):
        a=str(i*10+2)
        b=str(i*10+11)
        ws_result.merge_cells('A'+a+':A'+b)
    wb_result.save('最高成绩统计表.xlsx')

def test5():
    import openpyxl
    from collections import defaultdict

    def read_excel(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            movie_name, director, actors = row
            actors_list = actors.split(',')
            data.append((movie_name, actors_list))
        return data

    def find_best_actors(data, n):
        actor_movies = defaultdict(set)
        for movie_name, actors_list in data:
            for actor in actors_list:
                actor_movies[actor].add(movie_name)

        best_actors = sorted(actor_movies.items(), key=lambda x: len(x[1]), reverse=True)[:n]
        result = {actor: movies for actor, movies in best_actors}
        return result

    file_name = "电影导演演员.xlsx"
    n=int(input("输入一个大于或等于2的整数n:"))
    data = read_excel(file_name)
    best_actors = find_best_actors(data, n)
    for key,value in best_actors.items():
        print(f"{key}:{value}")

def test6():

    def generate():
        import pandas as pd
        import random
        import string

        file_names = ['E:/Python/Project/Test/Test5/test6/test'+str(i) + '.xlsx' for i in range(10)]

        for file_name in file_names:
            data = {'A': [random.randint(1, 100) for _ in range(10)],
                    'B': [random.choice(string.ascii_uppercase) for _ in range(10)],
                    'C': [random.uniform(1, 100) for _ in range(10)]}
            df = pd.DataFrame(data)
            df.to_excel(file_name, index=False,)

        print("已生成10个Excel文件并写入测试数据。")

    def manage(i):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, GradientFill, Color, colors
        workbook = openpyxl.load_workbook('E:/Python/Project/Test/Test5/test6/test'+str(i)+'.xlsx')
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color=colors.BLACK)
            for row in sheet.iter_rows(min_row=2):
                if row[0].row % 2 == 0:
                    for cell in row:
                        cell.font = Font(name='宋体', color="FF0000")
                    fill = GradientFill(stop=("0000FF", "FF0000"))
                    for cell in row:
                        cell.fill = fill
                else:
                    for cell in row:
                        cell.font = Font(name='宋体', color="ADD8E6")
                    #fill = PatternFill(patternType='solid', fgColor=Color('CCE5FF'))
                    #for cell in row:
                    #    cell.fill = fill
        workbook.save('E:/Python/Project/Test/Test5/test6/output'+str(i)+'.xlsx')

    generate()
    for i in range(10):
        manage(i)

def test8():
    import docx

    doc = docx.Document('test.docx')
    for para in doc.paragraphs:
        for run in para.runs:
            if run.font.color.rgb == docx.shared.RGBColor(255, 0, 0):
                print(f'Red text: {run.text}')

            if run.bold:
                print(f'Bold text: {run.text}')


if __name__=='__main__':
    test2()
    test4()
    test5()
    test6()
    test8()
