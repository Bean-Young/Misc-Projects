import openpyxl
import pandas as pd
import random
import os

header = ['ID', 'Group', 'Age', 'Gender', 'Address']

id_values_all = ['ID1', 'ID2', 'ID3', 'ID4', 'ID5','ID6','ID7']

def generate():
    save_path = r"E:\Python\Project\Test\Test5\test7"
    for i in range(1, 11):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for j, col in enumerate(header, start=1):
            sheet.cell(row=1, column=j).value = col
        num_rows_per_id=random.randint(1,10)
        data = []

        id_values=random.sample(id_values_all,5)

        for id_value in id_values:
            for k in range(1, num_rows_per_id + 1):
                group_value = random.randint(1, 10)
                age_value = random.randint(0, 100)
                gender_value = random.choice(['Male', 'Female'])
                address_value = f'Address {k}'
                data.append([id_value, group_value, age_value, gender_value, address_value])

            df = pd.DataFrame(data, columns=header)
            df = df.sort_values(by=['ID'])

        for l, row in df.iterrows():
            sheet.cell(row=l + 2, column=1).value = row['ID']
            sheet.cell(row=l + 2, column=2).value = row['Group']
            sheet.cell(row=l + 2, column=3).value = row['Age']
            sheet.cell(row=l + 2, column=4).value = row['Gender']
            sheet.cell(row=l + 2, column=5).value = row['Address']

        for lis in range(0,5):
            start_row = 2+lis*num_rows_per_id
            end_row = 1+(lis+1)*num_rows_per_id
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        file_name = f'file_{i}.xlsx'
        file_path = f"{save_path}\\{file_name}"
        workbook.save(file_path)


def work():
    folder_path = r"E:\Python\Project\Test\Test5\test7"
    save_path = r"E:\Python\Project\Test\Test5\test7"
    file_list = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
    merged_data = pd.DataFrame()

    for file in file_list:
        file_path = os.path.join(folder_path, file)
        df = pd.read_excel(file_path)
        value=''
        for i in range(0,len(df['ID'])):
            if df['ID'][i] not in id_values_all:
                df['ID'][i]=value
            else:
                value=df['ID'][i]
        #print(df)
        merged_data = pd.concat([merged_data, df],ignore_index=True)
    #print(merged_data)
    merged_data = merged_data.sort_values(by=['ID'],ignore_index=True)
    start_row_1 = 2
    end_row_1=0
    #print(len(merged_data['ID']))
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for j, col in enumerate(header, start=1):
        sheet.cell(row=1, column=j).value = col
    for l, row in merged_data.iterrows():
        sheet.cell(row=l + 2, column=1).value = row['ID']
        sheet.cell(row=l + 2, column=2).value = row['Group']
        sheet.cell(row=l + 2, column=3).value = row['Age']
        sheet.cell(row=l + 2, column=4).value = row['Gender']
        sheet.cell(row=l + 2, column=5).value = row['Address']
    for i in range(1, len(merged_data['ID'])):
        if not(merged_data['ID'][i]==merged_data['ID'][i-1]):
            end_row_1=i+1
            sheet.merge_cells(start_row=start_row_1, start_column=1, end_row=end_row_1, end_column=1)
            start_row_1=i+2
    #print(start_row_1)
    #print(i)
    sheet.merge_cells(start_row=start_row_1, start_column=1, end_row=i+2, end_column=1)
    file_name = 'merged_file.xlsx'
    file_path_1 = f"{save_path}\\{file_name}"
    workbook.save(file_path_1)

if __name__=='__main__':
    generate()
    work()
